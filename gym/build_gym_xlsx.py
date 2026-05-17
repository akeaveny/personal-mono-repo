"""Build gym.xlsx — master training file. Run from repo root."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "gym.xlsx"

PLAN = [
    (
        "Day 1 — Upper",
        [
            ("Incline Dumbbell Press", "4 x 8-12", "90 sec"),
            ("Cable Row", "4 x 10-12", "90 sec"),
            ("Dumbbell Shoulder Press", "3 x 10-12", "90 sec"),
            ("Lat Pulldown", "3 x 10-12", "90 sec"),
            ("Lateral Raises", "3 x 12-15", "60 sec"),
            ("Rear Delt Fly", "3 x 12-15", "60 sec"),
            ("Cable Crunches", "3 x 12-15", "60 sec"),
        ],
    ),
    (
        "Day 2 — Lower",
        [
            ("Leg Press", "4 x 10-12", "90 sec"),
            ("Bulgarian Split Squat", "3 x 10-12", "90 sec"),
            ("Lying Leg Curl", "3 x 10-12", "60 sec"),
            ("Leg Extension", "3 x 12-15", "60 sec"),
            ("Hip Thrust", "3 x 10-12", "90 sec"),
            ("Seated Calf Raise", "4 x 12-15", "60 sec"),
            ("Hanging Leg Raise", "3 x 10-15", "60 sec"),
        ],
    ),
    (
        "Day 3 — Push",
        [
            ("Flat Bench / Dumbbell Press", "4 x 8-12", "90 sec"),
            ("Low-to-High Cable Fly", "3 x 12-15", "60 sec"),
            ("Dips (lean forward)", "3 x 10-12", "90 sec"),
            ("Lateral Raises", "4 x 12-15", "60 sec"),
            ("Overhead Tricep Extension", "3 x 10-12", "60 sec"),
            ("Tricep Pushdown", "3 x 12-15", "60 sec"),
            ("Cable Crunches", "3 x 12-15", "60 sec"),
        ],
    ),
    (
        "Day 4 — Pull",
        [
            ("Lat Pulldown", "4 x 8-12", "90 sec"),
            ("Dumbbell Row", "4 x 10-12", "90 sec"),
            ("Face Pulls", "3 x 12-15", "60 sec"),
            ("Cable Lateral Raise", "3 x 12-15", "60 sec"),
            ("Incline Dumbbell Curl", "3 x 10-12", "60 sec"),
            ("Hammer Curl", "3 x 10-12", "60 sec"),
            ("Rear Delt Fly", "3 x 12-15", "60 sec"),
        ],
    ),
]

SCHEDULE = [
    ("Mon", "Upper", ""),
    ("Tue", "Easy ride (45-60 min, Zone 2)", "Recovery pace, conversational"),
    ("Wed", "Lower", ""),
    ("Thu", "Easy ride or rest", "Keep easy if legs are sore"),
    ("Fri", "Push", ""),
    ("Sat", "Pull", ""),
    ("Sun", "Long/hard ride (90 min+) or rest", "Big ride day, or rest if needed"),
]

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
DAY_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
DAY_FILL = PatternFill("solid", fgColor="2E75B6")
HDR_FONT = Font(name="Calibri", size=11, bold=True)
HDR_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")


def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_plan_sheet(wb):
    ws = wb.active
    ws.title = "Plan"

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Training Plan — Upper/Lower/Push/Pull")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    row = 3
    for day_title, exercises in PLAN:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        d = ws.cell(row=row, column=1, value=day_title)
        d.font = DAY_FONT
        d.fill = DAY_FILL
        d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        ws.cell(row=row, column=1, value="Exercise")
        ws.cell(row=row, column=2, value="Sets x Reps")
        ws.cell(row=row, column=3, value="Rest")
        style_header_row(ws, row, 3)
        row += 1

        for name, sets, rest in exercises:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=sets)
            ws.cell(row=row, column=3, value=rest)
            style_data_row(ws, row, 3)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"


def build_schedule_sheet(wb):
    ws = wb.create_sheet("Schedule")

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Weekly Schedule (Gym + Cycling)")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Day", "Session", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3)

    for i, (day, session, notes) in enumerate(SCHEDULE, start=4):
        ws.cell(row=i, column=1, value=day)
        ws.cell(row=i, column=2, value=session)
        ws.cell(row=i, column=3, value=notes)
        style_data_row(ws, i, 3)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 44
    ws.freeze_panes = "A4"


def parse_sets_reps(s):
    """'4 x 8-12' -> (sets=4, top_reps=12). '3 x 10-15' -> (3, 15)."""
    left, right = s.lower().split("x")
    sets = int(left.strip())
    rep_part = right.strip()
    top = int(rep_part.split("-")[1]) if "-" in rep_part else int(rep_part)
    return sets, top


def exercise_ref_rows():
    """Flatten PLAN into (exercise, day_label, planned_sets, top_reps)."""
    rows = []
    for day_title, exercises in PLAN:
        day_label = day_title.split("—")[0].strip()
        for name, sets_reps, _ in exercises:
            sets, top = parse_sets_reps(sets_reps)
            rows.append((name, day_label, sets, top))
    return rows


def build_log_sheet(wb):
    ws = wb.create_sheet("Log")

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="Workout Log — one row per set")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Date", "Day", "Exercise", "Set", "Weight (lb)", "Reps", "RPE", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    widths = {"A": 12, "B": 8, "C": 32, "D": 6, "E": 12, "F": 8, "G": 6, "H": 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(4, 504):
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"

    dv_day = DataValidation(
        type="list", formula1='"Day 1,Day 2,Day 3,Day 4"', allow_blank=True
    )
    ws.add_data_validation(dv_day)
    dv_day.add("B4:B1000")

    n = len(exercise_ref_rows())
    dv_ex = DataValidation(
        type="list", formula1=f"=Ref!$A$2:$A${n + 1}", allow_blank=True
    )
    ws.add_data_validation(dv_ex)
    dv_ex.add("C4:C1000")

    ws.freeze_panes = "A4"


def build_bodyweight_sheet(wb):
    ws = wb.create_sheet("Bodyweight")

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Bodyweight — weekly weigh-in")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Date", "Weight (lb)", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50

    for r in range(4, 304):
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2).number_format = "0.0"

    ws.freeze_panes = "A4"


def build_progress_sheet(wb):
    ws = wb.create_sheet("Progress")

    ws.merge_cells("A1:H1")
    t = ws.cell(
        row=1,
        column=1,
        value="Progress — top set + auto-suggestion (double progression, +5 lb)",
    )
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Exercise",
        "Last Logged",
        "Top Weight",
        "Min Reps @ Top",
        "Sets @ Top",
        "Target Sets",
        "Target Top Reps",
        "Suggestion",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    ref_rows = exercise_ref_rows()
    for i, (name, _, _, _) in enumerate(ref_rows, start=4):
        ws.cell(row=i, column=1, value=name)
        ws.cell(
            row=i,
            column=2,
            value=f'=IF(COUNTIF(Log!C:C,A{i})=0,"",MAXIFS(Log!A:A,Log!C:C,A{i}))',
        )
        ws.cell(
            row=i,
            column=3,
            value=f'=IF(B{i}="","",MAXIFS(Log!E:E,Log!C:C,A{i},Log!A:A,B{i}))',
        )
        ws.cell(
            row=i,
            column=4,
            value=f'=IF(B{i}="","",MINIFS(Log!F:F,Log!C:C,A{i},Log!A:A,B{i},Log!E:E,C{i}))',
        )
        ws.cell(
            row=i,
            column=5,
            value=f'=IF(B{i}="","",COUNTIFS(Log!C:C,A{i},Log!A:A,B{i},Log!E:E,C{i}))',
        )
        ws.cell(row=i, column=6, value=f"=VLOOKUP(A{i},Ref!A:D,3,FALSE)")
        ws.cell(row=i, column=7, value=f"=VLOOKUP(A{i},Ref!A:D,4,FALSE)")
        ws.cell(
            row=i,
            column=8,
            value=(
                f'=IF(B{i}="","No data — start with planned weight",'
                f'IF(C{i}=0,"Bodyweight — add reps or harder variation",'
                f'IF(AND(D{i}>=G{i},E{i}>=F{i}),'
                f'"Hit it — bump to "&(C{i}+5)&" lb",'
                f'"Stay at "&C{i}&" lb — add reps ("&D{i}&"/"&G{i}&")")))'
            ),
        )
        style_data_row(ws, i, len(headers))
        ws.cell(row=i, column=2).number_format = "yyyy-mm-dd"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 48
    ws.freeze_panes = "B4"


def build_ref_sheet(wb):
    ws = wb.create_sheet("Ref")

    headers = ["Exercise", "Day", "Planned Sets", "Top Reps"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    for i, row in enumerate(exercise_ref_rows(), start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val)
        style_data_row(ws, i, len(headers))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 11
    ws.sheet_state = "hidden"


def build_nutrition_sheet(wb):
    ws = wb.create_sheet("Nutrition")

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value="Daily Nutrition — manual entry from meal log")
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Date", "Calories", "Protein (g)", "Fat (g)", "Carbs (g)", "Binge?", "Notes / Triggers"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    widths = {"A": 12, "B": 10, "C": 12, "D": 10, "E": 10, "F": 8, "G": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    dv_binge = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True
    )
    ws.add_data_validation(dv_binge)
    dv_binge.add("F4:F500")

    for r in range(4, 504):
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"

    # Weekly summary row hint
    ws.cell(row=2, column=1, value="Target range:")
    ws.cell(row=2, column=2, value="2200-2600")
    ws.cell(row=2, column=3, value="150-180")
    for c in range(1, 4):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(name="Calibri", size=9, italic=True, color="808080")

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    build_plan_sheet(wb)
    build_schedule_sheet(wb)
    build_log_sheet(wb)
    build_bodyweight_sheet(wb)
    build_progress_sheet(wb)
    build_nutrition_sheet(wb)
    build_ref_sheet(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()

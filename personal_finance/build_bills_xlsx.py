"""Build/append formatted Bills tabs in personal_finance.xlsx.

Loads the existing workbook (if present) and appends two new dated tabs:
  - "Bills <YYYY-MM-DD>"
  - "Cards <YYYY-MM-DD>"
so the file accumulates monthly snapshots instead of overwriting.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "personal_finance.xlsx")
TODAY = date.today().isoformat()
BILLS_TAB = f"Bills {TODAY}"
CARDS_TAB = f"Cards {TODAY}"

if os.path.exists(XLSX_PATH) and os.path.getsize(XLSX_PATH) > 0:
    wb = openpyxl.load_workbook(XLSX_PATH)
    # If today's tabs already exist, suffix them so we never silently overwrite.
    suffix = 2
    while BILLS_TAB in wb.sheetnames or CARDS_TAB in wb.sheetnames:
        BILLS_TAB = f"Bills {TODAY} v{suffix}"
        CARDS_TAB = f"Cards {TODAY} v{suffix}"
        suffix += 1
else:
    wb = openpyxl.Workbook()
    # Drop the default empty sheet
    wb.remove(wb.active)

ws = wb.create_sheet(BILLS_TAB)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
SCOTIA_FILL = PatternFill("solid", fgColor="E2EFDA")
MOVE_FILL = PatternFill("solid", fgColor="FCE4D6")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws["A1"] = "Personal Finance - Monthly Bills & Card Strategy"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
ws.merge_cells("A1:G1")
ws["A2"] = f"Source: personal_finance.db (Feb-Apr 2026 actuals) | Generated {TODAY}"
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="595959")
ws.merge_cells("A2:G2")

headers = ["Bill", "Monthly $", "Currently Charged To", "Recommended Card", "Status", "Reward Rate", "Notes"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

sections = [
    ("HOUSING", [
        ("BLVD Beltline (rent)", 2053, "RBC PAD", "RBC PAD", "Keep", "n/a", "Landlords rarely accept CC. May 2026 = $2,003."),
    ]),
    ("TELECOM & UTILITIES", [
        ("Telus Mobility", 120, "RBC", "Scotia Momentum", "MOVE", "4%", "Volatile $86-$153/mo. Switch to BYOD plan to cut to ~$40 (Public Mobile)."),
        ("Moby Telecom", 47, "RBC", "Scotia Momentum", "MOVE", "4%", "Internet."),
        ("Enmax (electricity)", 71, "RBC PAD", "Scotia Momentum", "MOVE if allowed", "4%", "Confirm if Enmax accepts CC for PAD."),
        ("Metergy / EZ-Pay (utility)", 35, "RBC PAD", "Scotia Momentum", "MOVE if allowed", "4%", "Sub-meter billing."),
    ]),
    ("SUBSCRIPTIONS", [
        ("Apple.com/bill", 14, "RBC", "Scotia Momentum", "MOVE", "4%", "On board, undone."),
        ("ClassPass", 23, "RBC", "Scotia Momentum", "MOVE", "4%", "USD-billed Missoula MT."),
        ("Netflix", 8, "RBC WestJet", "RBC WestJet", "Keep", "1.5%", "Too small to bother moving."),
        ("YouTube Premium", 14, "Scotia Momentum", "Scotia Momentum", "OK", "4%", "Already correct."),
        ("Zwift", 26, "Scotia Momentum", "Scotia Momentum", "OK", "4%", "Already correct."),
    ]),
    ("INSURANCE", [
        ("Square One renters", 40, "Scotia Momentum", "Scotia Momentum", "OK", "4%", "Already correct."),
        ("Security National auto (annual)", 128, "Scotia Momentum", "Scotia Momentum", "OK", "4%", "$1,536/yr / 12. Paid Apr 23, 2026."),
    ]),
    ("BANK FEES", [
        ("RBC monthly account fee", 13, "RBC", "RBC", "Keep", "n/a", "Account fee, not moveable."),
    ]),
]

row = 5
for section, items in sections:
    cell = ws.cell(row=row, column=1, value=section)
    cell.fill = SUB_FILL
    cell.font = SUB_FONT
    cell.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = SUB_FILL
    row += 1
    for bill, amt, current, recommended, status, reward, notes in items:
        ws.cell(row=row, column=1, value=bill).font = BODY_FONT
        amt_cell = ws.cell(row=row, column=2, value=amt)
        amt_cell.number_format = '"$"#,##0.00'
        amt_cell.alignment = RIGHT
        amt_cell.font = BODY_FONT
        ws.cell(row=row, column=3, value=current).font = BODY_FONT
        ws.cell(row=row, column=4, value=recommended).font = BODY_FONT
        status_cell = ws.cell(row=row, column=5, value=status)
        status_cell.alignment = CENTER
        status_cell.font = Font(name="Calibri", size=10, bold=True)
        if status.startswith("MOVE"):
            status_cell.fill = MOVE_FILL
        elif status == "OK":
            status_cell.fill = SCOTIA_FILL
        ws.cell(row=row, column=6, value=reward).alignment = CENTER
        ws.cell(row=row, column=6).font = BODY_FONT
        notes_cell = ws.cell(row=row, column=7, value=notes)
        notes_cell.alignment = WRAP
        notes_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

total_row = row
ws.cell(row=total_row, column=1, value="TOTAL (excl. rent)").font = TOTAL_FONT
ws.cell(row=total_row, column=1).fill = TOTAL_FILL
total_excl_rent = sum(amt for _, items in sections for bill, amt, *_ in items if "rent" not in bill.lower())
total_cell = ws.cell(row=total_row, column=2, value=total_excl_rent)
total_cell.number_format = '"$"#,##0.00'
total_cell.font = TOTAL_FONT
total_cell.fill = TOTAL_FILL
total_cell.alignment = RIGHT
ws.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=7)
for col in range(3, 8):
    ws.cell(row=total_row, column=col).fill = TOTAL_FILL
for col in range(1, 8):
    ws.cell(row=total_row, column=col).border = BORDER
row += 1

all_in_row = row
ws.cell(row=all_in_row, column=1, value="ALL-IN MONTHLY (incl. rent)").font = TOTAL_FONT
ws.cell(row=all_in_row, column=1).fill = TOTAL_FILL
total_all = sum(amt for _, items in sections for _, amt, *_ in items)
all_cell = ws.cell(row=all_in_row, column=2, value=total_all)
all_cell.number_format = '"$"#,##0.00'
all_cell.font = TOTAL_FONT
all_cell.fill = TOTAL_FILL
all_cell.alignment = RIGHT
ws.merge_cells(start_row=all_in_row, start_column=3, end_row=all_in_row, end_column=7)
for col in range(3, 8):
    ws.cell(row=all_in_row, column=col).fill = TOTAL_FILL
for col in range(1, 8):
    ws.cell(row=all_in_row, column=col).border = BORDER
row += 2

ws.cell(row=row, column=1, value="REWARD UPLIFT FROM MOVING BILLS TO SCOTIA").font = SUB_FONT
ws.cell(row=row, column=1).fill = SUB_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
for col in range(1, 8):
    ws.cell(row=row, column=col).fill = SUB_FILL
row += 1

uplift_headers = ["Bill", "Monthly $", "Scotia 4%/mo", "RBC 1.5%/mo", "Net Gain/mo"]
for col, h in enumerate(uplift_headers, 1):
    c = ws.cell(row=row, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
row += 1

moveable = [
    ("Telus Mobility", 120),
    ("Moby Telecom", 47),
    ("Enmax (if accepts CC)", 71),
    ("Metergy (if accepts CC)", 35),
    ("Apple", 14),
    ("ClassPass", 23),
]
move_total = 0
for bill, amt in moveable:
    ws.cell(row=row, column=1, value=bill).font = BODY_FONT
    c = ws.cell(row=row, column=2, value=amt)
    c.number_format = '"$"#,##0.00'
    c.alignment = RIGHT
    c = ws.cell(row=row, column=3, value=amt * 0.04)
    c.number_format = '"$"#,##0.00'
    c.alignment = RIGHT
    c = ws.cell(row=row, column=4, value=amt * 0.015)
    c.number_format = '"$"#,##0.00'
    c.alignment = RIGHT
    c = ws.cell(row=row, column=5, value=amt * 0.025)
    c.number_format = '"$"#,##0.00'
    c.alignment = RIGHT
    c.font = Font(name="Calibri", size=10, bold=True, color="00B050")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = BORDER
    move_total += amt
    row += 1

ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
ws.cell(row=row, column=1).fill = TOTAL_FILL
for col, val in [(2, move_total), (3, move_total * 0.04), (4, move_total * 0.015), (5, move_total * 0.025)]:
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = '"$"#,##0.00'
    c.font = TOTAL_FONT
    c.fill = TOTAL_FILL
    c.alignment = RIGHT
    c.border = BORDER
ws.cell(row=row, column=1).border = BORDER
row += 1
annual_gain = move_total * 0.025 * 12
ws.cell(row=row, column=1, value=f"Annual net gain: ${annual_gain:.2f}  (Scotia annual fee = $120, so net keep = ${annual_gain - 120:.2f}/yr)").font = Font(name="Calibri", size=10, italic=True, bold=True, color="00B050")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

ws.cell(row=row, column=1, value="ACTION ITEMS - UPDATE BILLING CARD").font = SUB_FONT
ws.cell(row=row, column=1).fill = SUB_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
for col in range(1, 8):
    ws.cell(row=row, column=col).fill = SUB_FILL
row += 1
actions = [
    "1. Telus.com -> Account -> Payment Method -> swap to Scotia ...7582",
    "2. Apple ID (appleid.apple.com) -> Payment & Shipping -> swap to Scotia",
    "3. Moby account portal -> billing -> swap to Scotia",
    "4. ClassPass app -> Payment -> swap to Scotia",
    "5. Enmax: call/check if CC PAD allowed; if yes -> Scotia",
    "6. Metergy: call/check if CC PAD allowed; if yes -> Scotia",
    "7. Once on Scotia, request $1,208.90 credit refund from Scotia",
    "8. Confirm Eau Claire $221 (Apr 14) - annual or monthly?",
]
for a in actions:
    ws.cell(row=row, column=1, value=a).font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

widths = {1: 36, 2: 12, 3: 22, 4: 22, 5: 16, 6: 12, 7: 50}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A5"

ws2 = wb.create_sheet(CARDS_TAB)
ws2["A1"] = "Credit Card Assignments"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
ws2.merge_cells("A1:E1")

card_headers = ["Card", "Limit", "Use For", "Reward Rate", "Notes"]
for col, h in enumerate(card_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

cards = [
    ("Scotia Momentum VISA Infinite (...7582)", "$21,500", "ALL recurring bills, groceries, gas, transit", "4% bills/groceries | 2% gas/transit", "Annual fee $120 - justified by ~$300+/yr rewards"),
    ("RBC WestJet World Elite MC (...5104)", "$18,500", "Dining, UberEats, Amazon, travel, everything else", "~1.5% WestJet $ + travel insurance", "Default fallback card"),
    ("RBC Cash Back MC (...6035)", "$13,000", "DON'T USE - sock drawer", "25.99% APR - useless rate", "Keep open for credit limit / utilization only"),
]
r = 4
for card, limit, use, reward, notes in cards:
    ws2.cell(row=r, column=1, value=card).font = Font(name="Calibri", size=10, bold=True)
    ws2.cell(row=r, column=2, value=limit).alignment = CENTER
    ws2.cell(row=r, column=3, value=use).alignment = WRAP
    ws2.cell(row=r, column=4, value=reward).alignment = WRAP
    ws2.cell(row=r, column=5, value=notes).alignment = WRAP
    for col in range(1, 6):
        ws2.cell(row=r, column=col).border = BORDER
        if col != 1:
            ws2.cell(row=r, column=col).font = BODY_FONT
    ws2.row_dimensions[r].height = 45
    r += 1

widths2 = {1: 38, 2: 12, 3: 40, 4: 32, 5: 40}
for col, w in widths2.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

wb.save(XLSX_PATH)
print(f"Saved. Sheets now: {wb.sheetnames}")
print(f"Added: '{BILLS_TAB}', '{CARDS_TAB}'")
